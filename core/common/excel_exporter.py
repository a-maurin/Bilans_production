# Copyright (C) 2026 Aguirre MAURIN
#
# Ce programme est un logiciel libre : vous pouvez le redistribuer et/ou le modifier
# selon les termes de la Licence Publique Générale GNU (GPL) telle que publiée par
# la Free Software Foundation, version 3 de la licence, ou (à votre choix) toute version ultérieure.
#
# Ce programme est distribué dans l'espoir qu'il sera utile, mais SANS AUCUNE GARANTIE ;
# sans même la garantie implicite de QUALITÉ MARCHANDE ou D'ADÉQUATION À UN USAGE PARTICULIER.
# Voir la Licence Publique Générale GNU pour plus de détails.
#
# CONDITIONS SUPPLÉMENTAIRES D'ATTRIBUTION (SECTION 7(b) DE LA GPL v3) :
# Conformément à la section 7(b) DE LA GPL v3, vous devez expressément conserver
# intactes et lisibles toutes les mentions d'auteur, notices de copyright et la présente
# clause dans chaque fichier source ou interface utilisateur redistribué. Toute version modifiée
# doit clairement indiquer qu'elle a été altérée et ne doit en aucun cas supprimer le nom
# de l'auteur original (Aguirre MAURIN).

"""
========================================================================================
MODULE : EXPORTATEUR DE DONNÉES EN CLASSEURS EXCEL (`excel_exporter.py`)
========================================================================================
Ce module s'occupe de la création et de la mise en forme automatique des fichiers Excel (.xlsx).

Fonctionnalités principales :
  1. Application de la charte graphique OFB (en-têtes bleus `#003366`, polices Arial, bordures grises).
  2. Ajustement automatique de la largeur des colonnes selon la longueur du contenu.
  3. Génération de classeurs multi-onglets (Synthèse Régionale + 1 onglet par Département).
========================================================================================
"""

import logging
from pathlib import Path
from typing import Any, List, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ========================================================================================
# MISE EN FORME DES FEUILLES DE CALCUL (STYLES OFB ET AUTO-FIT)
# ========================================================================================

def _format_worksheet(ws: Any) -> None:
    """Applique la charte graphique OFB et ajuste automatiquement la largeur des colonnes."""
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    if hasattr(ws.views, "sheetView") and ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    # 1. En-têtes (fond bleu roi, texte blanc gras, centré)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 2. Données et bordures (nombres à droite, textes à gauche)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 3. Calcul dynamique de la largeur optimale des colonnes
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


# ========================================================================================
# EXPORT DE LA SYNTHESE REGIONALE MULTI-ONGLETS
# ========================================================================================

def export_bilan_excel(
    out_dir: Path,
    df_detail: pd.DataFrame,
    echelle: str = "region",
    code: Optional[str] = None,
    depts: Optional[List[str]] = None,
    df_ratio: Optional[pd.DataFrame] = None,
    filename: Optional[str] = None
) -> Optional[Path]:
    """Exporte un bilan d'analyse complet en classeur Excel adaptatif selon l'échelle.

    - Échelle Région : Synthese_Regionale + Dept_XX (par département) + Donnees_Brutes
    - Échelle Département : Synthese_Dept_XX + Donnees_Brutes
    """
    if df_detail is None or df_detail.empty:
        logger.warning("df_detail est vide ou None, annulation de l'export Excel.")
        return None

    if not filename:
        code_str = str(code) if code else "global"
        filename = f"bilan_export_{echelle}_{code_str}.xlsx"

    xlsx_path = out_dir / filename
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            if echelle == "departement":
                sheet_main = f"Synthese_Dept_{code}"[:31] if code else "Synthese_Departement"
                if df_ratio is not None and not df_ratio.empty:
                    df_ratio.to_excel(writer, sheet_name=sheet_main, index=False)
                else:
                    cols = [c for c in ["nb_operations", "nb_localisations", "nb_pej", "nb_pa", "nb_pve"] if c in df_detail.columns]
                    group_cols = [c for c in ["domaine", "theme"] if c in df_detail.columns]
                    if group_cols and cols:
                        agg_df = df_detail.groupby(group_cols)[cols].sum().reset_index()
                        agg_df.to_excel(writer, sheet_name=sheet_main, index=False)
                    else:
                        df_detail.to_excel(writer, sheet_name=sheet_main, index=False)
                _format_worksheet(writer.sheets[sheet_main])

                # Onglet Données Brutes
                df_detail.to_excel(writer, sheet_name="Donnees_Brutes", index=False)
                _format_worksheet(writer.sheets["Donnees_Brutes"])

            else:
                # Échelle Régionale / par défaut
                sheet_main = "Synthese_Regionale"
                if df_ratio is not None and not df_ratio.empty:
                    df_ratio.to_excel(writer, sheet_name=sheet_main, index=False)
                else:
                    cols = [c for c in ["nb_operations", "nb_localisations", "nb_pej", "nb_pa", "nb_pve"] if c in df_detail.columns]
                    group_cols = [c for c in ["domaine", "theme"] if c in df_detail.columns]
                    if group_cols and cols:
                        agg_region = df_detail.groupby(group_cols)[cols].sum().reset_index()
                        agg_region.to_excel(writer, sheet_name=sheet_main, index=False)
                    else:
                        df_detail.to_excel(writer, sheet_name=sheet_main, index=False)
                _format_worksheet(writer.sheets[sheet_main])

                # Onglets départementaux
                if "departement" in df_detail.columns:
                    df_detail["departement"] = df_detail["departement"].astype(str)
                    target_depts = depts or [str(d) for d in df_detail["departement"].dropna().unique()]
                    for dept in target_depts:
                        dept_str = str(dept)
                        df_dept = df_detail[df_detail["departement"] == dept_str]
                        sheet_title = f"Dept_{dept_str}"[:31]
                        if not df_dept.empty:
                            df_dept.to_excel(writer, sheet_name=sheet_title, index=False)
                        else:
                            pd.DataFrame({"Message": ["Aucune donnée disponible pour ce département"]}).to_excel(
                                writer, sheet_name=sheet_title, index=False
                            )
                        _format_worksheet(writer.sheets[sheet_title])

                # Onglet Données Brutes
                df_detail.to_excel(writer, sheet_name="Donnees_Brutes", index=False)
                _format_worksheet(writer.sheets["Donnees_Brutes"])

        logger.info(f"Classeur Excel formaté généré avec succès : {xlsx_path}")
        return xlsx_path

    except Exception as e:
        logger.error(f"Erreur lors de la génération du classeur Excel {xlsx_path} : {e}")
        return None


def export_synthese_region_excel(
    out_dir: Path,
    df_detail: pd.DataFrame,
    depts: List[str],
    df_ratio: Optional[pd.DataFrame] = None,
    filename: str = "Synthese_Region.xlsx"
) -> Optional[Path]:
    """Exporte la synthèse régionale et les détails départementaux dans un classeur Excel."""
    return export_bilan_excel(
        out_dir=out_dir,
        df_detail=df_detail,
        echelle="region",
        depts=depts,
        df_ratio=df_ratio,
        filename=filename
    )


# ========================================================================================
# INTERFACE LIGNE DE COMMANDE (CLI)
# ========================================================================================

def load_input_file(input_path: Path) -> pd.DataFrame:
    """Charge un fichier CSV ou Excel en DataFrame pandas."""
    if not input_path.exists():
        raise FileNotFoundError(f"Le fichier spécifié n'existe pas : {input_path}")
    
    ext = input_path.suffix.lower()
    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(input_path)
    elif ext in [".csv", ".txt"]:
        try:
            return pd.read_csv(input_path, sep=";", encoding="utf-8")
        except Exception:
            return pd.read_csv(input_path, sep=",", encoding="utf-8")
    else:
        raise ValueError(f"Format de fichier non pris en charge : {ext} (formats acceptés : .csv, .xlsx, .xls)")


def main(argv: Optional[List[str]] = None) -> int:
    """Point d'entrée CLI pour la génération de classeurs Excel."""
    import argparse
    import sys

    parser = argparse.ArgumentParser(
        description="Générateur de classeurs Excel de synthèse OFBilan."
    )
    parser.add_argument(
        "-i", "--input", required=True, type=Path, help="Chemin du fichier source (CSV ou Excel)"
    )
    parser.add_argument(
        "-o", "--out-dir", type=Path, default=Path("."), help="Dossier de destination (défaut: .)"
    )
    parser.add_argument(
        "-f", "--filename", type=str, default="Synthese_Region.xlsx", help="Nom du fichier Excel généré (défaut: Synthese_Region.xlsx)"
    )
    parser.add_argument(
        "-e", "--echelle", choices=["region", "departement"], default="region", help="Échelle spatiale (défaut: region)"
    )
    parser.add_argument(
        "-c", "--code", type=str, default=None, help="Code géographique (ex: 21 ou r27). Requis si --echelle departement."
    )
    parser.add_argument(
        "--depts", nargs="+", default=None, help="Liste explicite des codes départements (ex: --depts 21 25 39)"
    )

    args = parser.parse_args(argv)

    if args.echelle == "departement" and not args.code:
        parser.error("L'argument --code est obligatoire lorsque --echelle est réglé sur 'departement'.")

    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Chargement des données depuis {args.input}...")
    df_detail = load_input_file(args.input)

    if args.echelle == "departement":
        depts = [str(args.code)]
    elif args.depts:
        depts = [str(d) for d in args.depts]
    else:
        if "departement" in df_detail.columns:
            depts = [str(d) for d in df_detail["departement"].dropna().unique()]
        else:
            depts = []

    res_path = export_synthese_region_excel(
        out_dir=out_dir,
        df_detail=df_detail,
        depts=depts,
        filename=args.filename
    )

    if res_path:
        print(f"✅ Classeur Excel généré avec succès : {res_path.resolve()}")
        return 0
    else:
        print("❌ Échec de la génération du classeur Excel.")
        return 1


if __name__ == "__main__":
    import sys
    sys.exit(main())


